"""
TODO: Custom django admin command to load UO courses into the database 

2023-05-25 - Josh Sawyer : Created command to add courses to the database

"""

from django.core.management.base import BaseCommand, CommandError
from forecast.models import Course, Keyword
from users.models import User
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

class Command(BaseCommand):
    help = """Imports courses from an Excel file.
          The excel file must contain the following fields: 
          id, subject, number, credits, term, annual, name, required, and keywords
          """
    
    def add_arguments(self, parser):
        parser.add_argument("file_path", type=str, help="Path to excel file")
    
    def handle(self, *args, **options):
        # Get the path to the excel sheet (that was provided by the user)
        file = options['file_path']
        
        try:
            workbook = load_workbook(file)
            sheet = workbook.active
          
            ID = 0
            NAME = 7
            SUBJECT = 1
            NUMBER = 2
            PREREQ = 3
            CREDITS = 4
            TERM = 5
            ANNUAL = 6
            KEYWORDS = 9
            
            courses = []

            # Go through each row in the excel sheet and get the necessary information to add
            # to the course object
            for row in sheet.iter_rows(min_row=2, values_only=True):
                
                # Check if there's a keyword associated with this course
                if (row[KEYWORDS] != "~" and row[KEYWORDS] is not None):
                    keywords = row[KEYWORDS].split(', ') # All keywords should be separated with ", "
                    for keyword in keywords:
                        # Check if this keyword has been added into the database
                        if not Keyword.objects.filter(keyword=keyword).exists():
                            Keyword.objects.create(keyword=keyword) # Add keyword to the database
                            # self.stdout.write(f"Keyword [{keyword}] added to database.")
                        else:
                            self.stdout.write(f"Keyword [{keyword}] already exists, skipping.")
                
                # Check to see if the row has an ID
                if row[ID] is not None:
                    # Check if the course already exists
                    if Course.objects.filter(id=row[ID]).exists():
                        self.stdout.write(f"Course already exists, skipping.")
                        continue
                
                terms = {'F': False, 'W': False, 'S': False}
                for term in terms: # Get all the terms the course is offered
                    if (row[TERM] is not None):
                        if term in row[TERM]:
                            terms[term] = True
                
                every_year = False 
                if type(row[ANNUAL]) is type(bool): # Check if it's offered every year
                    every_year = True

                course = Course( # Create the course object
                    id = int(row[ID]),
                    name = str(row[NAME]),
                    subject = row[SUBJECT],
                    number = row[NUMBER],
                    credits = row[CREDITS],
                    fall = terms['F'],
                    winter = terms['W'],
                    spring = terms['S'],
                    every_year = every_year
                )
                
                
              
                courses.append(course) # Append it to a list that'll be used to bulk create
            
            Course.objects.bulk_create(courses)
            
            self.stdout.write(f"Courses added, now adding prereqs.")
            
            # Courses have now been created, now add prereqs and tie keywords in
            for row in sheet.iter_rows(min_row=2, values_only=True):
                course = Course.objects.get(id=row[ID]) # Get the course associated with this row

                if row[KEYWORDS] is not None and row[KEYWORDS] != "~":
                    keywords = row[KEYWORDS].split(', ')
                    for keyword in keywords:
                        if not course.has_kw.filter(keyword=keyword).exists():
                            keyword_obj = Keyword.objects.get(keyword=keyword)
                            course.has_kw.add(keyword_obj)
                            # self.stdout.write(f"Keyword [{keyword}] added to the course [{course}].")
                        else:
                            self.stdout.write(f"Course [{course}] already has the keyword [{keyword}].")
                
                prereqs = row[PREREQ]
                if prereqs is not None:
                    prereq_ids = []
                    if type(prereqs) is str: # In the case of multiple prereqs
                        prereqs_list = row[PREREQ].split(', ')
                        for prereq in prereqs_list:
                            prereq_ids.append(int(prereq))
                    else:
                        if (int(prereqs) != 0):
                            prereq_ids.append(int(prereqs))
                    
                    for prereq_id in prereq_ids: 
                        try:
                            if not course.has_prereq.filter(id=prereq_id).exists():
                                prereq_course = Course.objects.get(id=prereq_id)
                                course.has_prereq.add(prereq_course)
                            else:
                                self.stdout.write(f"Course already has this prereq.")
                            
                        except Course.DoesNotExist:
                            self.stderr.write(f"Course doesn't exist: prereq_id [{prereq_id}]")
            
            self.stdout.write(f"Prereqs added, course import now complete.")
            
        except InvalidFileException:
            self.stderr.write("Invalid Excel file.")


        try:
            user = User.objects.create_user(
                username = "admin",
                email = "test@test.com",
                password = "admin",
                first_name = "Puddles",
                last_name = "Duck"
            )
            user.save()
            self.stdout.write("Created default user")

        except:
            self.stderr.write("Unable to create default user.")
